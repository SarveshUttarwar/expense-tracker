import io
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def generate_excel_report(expenses: list):
    """Generate a professionally formatted Excel report with headers, styling, and totals."""
    df = pd.DataFrame(expenses)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Expenses')
        ws = writer.sheets['Expenses']

        # --- Header styling ---
        header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- Body styling ---
        body_font = Font(name='Calibri', size=11)
        body_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        expense_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
        saving_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), 2):
            for cell in row:
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = thin_border

            # Color-code by type column (index 1 = "Type")
            type_cell = ws.cell(row=row_idx, column=2)
            if type_cell.value and str(type_cell.value).lower() == 'saving':
                for cell in row:
                    cell.fill = saving_fill
            elif type_cell.value and str(type_cell.value).lower() == 'expense':
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = expense_fill
                else:
                    for cell in row:
                        cell.fill = alt_fill

        # --- Totals row ---
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name='Calibri', bold=True, size=12)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')

        # Amount column is column 4
        total_amount = sum(float(r.get('Amount', 0) or 0) for r in expenses)
        amount_cell = ws.cell(row=total_row, column=4, value=total_amount)
        amount_cell.font = Font(name='Calibri', bold=True, size=12, color='4F46E5')
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = Alignment(horizontal='center')

        for cell_idx in range(1, ws.max_column + 1):
            ws.cell(row=total_row, column=cell_idx).border = Border(
                top=Side(style='double', color='4F46E5'),
                bottom=Side(style='double', color='4F46E5'),
            )

        # --- Auto-fit column widths ---
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    output.seek(0)
    return output


def generate_pdf_report(expenses: list):
    """Generate a professionally formatted PDF report with table, summary, and page numbers."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=50,
        bottomMargin=50,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1E1B4B'),
        spaceAfter=6,
        fontName='Helvetica-Bold',
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=20,
        fontName='Helvetica',
    )
    summary_label_style = ParagraphStyle(
        'SummaryLabel',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#6B7280'),
        fontName='Helvetica',
    )
    summary_value_style = ParagraphStyle(
        'SummaryValue',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1E1B4B'),
        fontName='Helvetica-Bold',
    )

    elements = []

    # --- Title ---
    elements.append(Paragraph("Expense Report", title_style))
    now = datetime.now()
    elements.append(Paragraph(f"Generated on {now.strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))

    # --- Summary cards ---
    total_expenses = sum(float(r.get('Amount', 0) or 0) for r in expenses if str(r.get('Type', '')).lower() == 'expense')
    total_savings = sum(float(r.get('Amount', 0) or 0) for r in expenses if str(r.get('Type', '')).lower() == 'saving')
    total_transactions = len(expenses)

    # Category breakdown
    cat_totals = {}
    for r in expenses:
        if str(r.get('Type', '')).lower() == 'expense':
            cat = r.get('Category', 'Unknown') or 'Unknown'
            cat_totals[cat] = cat_totals.get(cat, 0) + float(r.get('Amount', 0) or 0)

    summary_data = [
        [
            Paragraph("Total Expenses", summary_label_style),
            Paragraph("Total Savings", summary_label_style),
            Paragraph("Transactions", summary_label_style),
        ],
        [
            Paragraph(f"₹{total_expenses:,.2f}", summary_value_style),
            Paragraph(f"₹{total_savings:,.2f}", summary_value_style),
            Paragraph(str(total_transactions), summary_value_style),
        ],
    ]

    summary_table = Table(summary_data, colWidths=[170, 170, 170])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('RIGHTPADDING', (0, 0), (-1, -1), 14),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # --- Category breakdown ---
    if cat_totals:
        elements.append(Paragraph("Category Breakdown", ParagraphStyle(
            'CatTitle', parent=styles['Heading2'], fontSize=13,
            textColor=colors.HexColor('#1E1B4B'), spaceAfter=8,
            fontName='Helvetica-Bold',
        )))

        cat_data = [["Category", "Total (₹)"]]
        for cat_name, cat_val in sorted(cat_totals.items(), key=lambda x: x[1], reverse=True):
            cat_data.append([cat_name, f"₹{cat_val:,.2f}"])

        cat_table = Table(cat_data, colWidths=[300, 200])
        cat_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        for i in range(1, len(cat_data)):
            if i % 2 == 0:
                cat_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
        cat_table.setStyle(TableStyle(cat_style))
        elements.append(cat_table)
        elements.append(Spacer(1, 20))

    # --- Transaction table ---
    elements.append(Paragraph("All Transactions", ParagraphStyle(
        'TxTitle', parent=styles['Heading2'], fontSize=13,
        textColor=colors.HexColor('#1E1B4B'), spaceAfter=8,
        fontName='Helvetica-Bold',
    )))

    table_data = [["Date", "Type", "Category", "Amount (₹)", "Description"]]
    for exp in expenses:
        desc = str(exp.get('Description', '') or '')
        if len(desc) > 45:
            desc = desc[:42] + "..."
        table_data.append([
            str(exp.get('Date', '')),
            str(exp.get('Type', '')).capitalize(),
            str(exp.get('Category', '') or 'N/A'),
            f"₹{float(exp.get('Amount', 0) or 0):,.2f}",
            desc,
        ])

    col_widths = [70, 55, 90, 80, 230]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
    ]

    # Alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))

    main_table.setStyle(TableStyle(table_style))
    elements.append(main_table)

    # --- Build with page numbers ---
    def add_page_number(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 8)
        canvas_obj.setFillColor(colors.HexColor('#9CA3AF'))
        page_num = canvas_obj.getPageNumber()
        canvas_obj.drawRightString(
            letter[0] - 40, 30,
            f"Page {page_num}"
        )
        canvas_obj.drawString(40, 30, "Expense Tracker Report")
        canvas_obj.restoreState()

    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    output.seek(0)
    return output
